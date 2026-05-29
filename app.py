import calendar
import csv
import io
import os
import sqlite3
import tempfile
from datetime import date, datetime, timedelta
from functools import wraps

import jwt
import mysql.connector
from authlib.integrations.flask_client import OAuth
from dotenv import load_dotenv
from flask import (
    Flask,
    Response,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    session,
    url_for,
)
from openpyxl import Workbook, load_workbook
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

load_dotenv()

app = Flask(__name__, template_folder="app/templates", static_folder="app/static")
app.config["SECRET_KEY"] = os.getenv("SECRET_KEY", "dev-secret-key")
app.config["UPLOAD_FOLDER"] = os.path.join(app.static_folder, "uploads")
app.config["MAX_CONTENT_LENGTH"] = 2 * 1024 * 1024
JWT_SECRET = os.getenv("JWT_SECRET", app.config["SECRET_KEY"])
DB_DRIVER = os.getenv("DB_DRIVER", "mysql").lower()
SQLITE_FALLBACK = os.getenv("SQLITE_FALLBACK", "true").lower() == "true"
SQLITE_PATH = os.getenv("SQLITE_PATH") or os.path.join(tempfile.gettempdir(), "finance_tracker.sqlite3")
ACTIVE_DB_DRIVER = DB_DRIVER
SQLITE_READY = False

os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)

oauth = OAuth(app)
google = oauth.register(
    name="google",
    client_id=os.getenv("GOOGLE_CLIENT_ID"),
    client_secret=os.getenv("GOOGLE_CLIENT_SECRET"),
    server_metadata_url="https://accounts.google.com/.well-known/openid-configuration",
    client_kwargs={"scope": "openid email profile"},
)

CATEGORIES = [
    "Food",
    "Transportation",
    "Shopping",
    "Education",
    "Healthcare",
    "Entertainment",
    "Salary",
    "Other",
]

EXCEL_COLUMNS = ["title", "amount", "category", "type", "transaction_date", "notes"]


class DatabaseUnavailable(RuntimeError):
    pass


def mysql_db():
    return mysql.connector.connect(
        host=os.getenv("MYSQL_HOST", "localhost"),
        port=int(os.getenv("MYSQL_PORT", "3306")),
        user=os.getenv("MYSQL_USER", "root"),
        password=os.getenv("MYSQL_PASSWORD", ""),
        database=os.getenv("MYSQL_DATABASE", "finance_tracker"),
    )


def init_sqlite():
    global SQLITE_READY
    if SQLITE_READY:
        return
    os.makedirs(os.path.dirname(SQLITE_PATH), exist_ok=True)
    conn = sqlite3.connect(SQLITE_PATH)
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS users (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          name TEXT NOT NULL,
          email TEXT NOT NULL UNIQUE,
          password TEXT NOT NULL,
          profile_photo TEXT,
          created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS transactions (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          user_id INTEGER NOT NULL,
          title TEXT NOT NULL,
          amount REAL NOT NULL,
          category TEXT NOT NULL,
          type TEXT NOT NULL,
          notes TEXT,
          transaction_date TEXT NOT NULL,
          created_at TEXT DEFAULT CURRENT_TIMESTAMP,
          FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
        );
        CREATE TABLE IF NOT EXISTS budgets (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          user_id INTEGER NOT NULL,
          monthly_budget REAL NOT NULL,
          month INTEGER NOT NULL,
          year INTEGER NOT NULL,
          created_at TEXT DEFAULT CURRENT_TIMESTAMP,
          UNIQUE(user_id, month, year),
          FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
        );
        CREATE TABLE IF NOT EXISTS alerts (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          user_id INTEGER NOT NULL,
          message TEXT NOT NULL,
          alert_type TEXT NOT NULL,
          created_at TEXT DEFAULT CURRENT_TIMESTAMP,
          FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
        );
        """
    )
    user_count = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
    if user_count == 0:
        password = generate_password_hash("password123")
        cur = conn.execute(
            "INSERT INTO users (name, email, password) VALUES (?,?,?)",
            ("Demo User", "demo@example.com", password),
        )
        user_id = cur.lastrowid
        today = date.today()
        conn.execute(
            "INSERT INTO budgets (user_id, monthly_budget, month, year) VALUES (?,?,?,?)",
            (user_id, 2500.00, today.month, today.year),
        )
        rows = [
            ("Monthly Salary", 4200.00, "Salary", "Income", "Primary job", today - timedelta(days=20)),
            ("Groceries", 340.25, "Food", "Expense", "Weekly grocery run", today - timedelta(days=18)),
            ("Metro card", 65.00, "Transportation", "Expense", "Commute pass", today - timedelta(days=15)),
            ("Online course", 199.00, "Education", "Expense", "Skill upgrade", today - timedelta(days=12)),
            ("Family dinner", 120.00, "Entertainment", "Expense", "Weekend outing", today - timedelta(days=8)),
            ("Medicine", 74.50, "Healthcare", "Expense", "Pharmacy", today - timedelta(days=6)),
            ("Freelance Payment", 850.00, "Salary", "Income", "Side project", today - timedelta(days=4)),
            ("New shoes", 150.00, "Shopping", "Expense", "Seasonal purchase", today - timedelta(days=2)),
        ]
        conn.executemany(
            """INSERT INTO transactions (user_id, title, amount, category, type, notes, transaction_date)
               VALUES (?,?,?,?,?,?,?)""",
            [(user_id, title, amount, category, tx_type, notes, tx_date.isoformat()) for title, amount, category, tx_type, notes, tx_date in rows],
        )
        conn.execute(
            "INSERT INTO alerts (user_id, message, alert_type) VALUES (?,?,?)",
            (user_id, "Warning! You have used 80% of your budget.", "warning"),
        )
    conn.commit()
    conn.close()
    SQLITE_READY = True


def sqlite_db():
    init_sqlite()
    conn = sqlite3.connect(SQLITE_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def db():
    global ACTIVE_DB_DRIVER
    if ACTIVE_DB_DRIVER == "sqlite":
        return sqlite_db(), "sqlite"
    try:
        return mysql_db(), "mysql"
    except mysql.connector.Error as exc:
        if SQLITE_FALLBACK:
            ACTIVE_DB_DRIVER = "sqlite"
            return sqlite_db(), "sqlite"
        raise DatabaseUnavailable(str(exc)) from exc


def normalize_sql(sql, driver):
    if driver != "sqlite":
        return sql
    replacements = {
        "%s": "?",
        "DATE_FORMAT(transaction_date, '%Y-%m')": "strftime('%Y-%m', transaction_date)",
        "DATE_FORMAT(transaction_date, '%Y-%m-%d')": "strftime('%Y-%m-%d', transaction_date)",
        "DATE_FORMAT(transaction_date, '%x-W%v')": "strftime('%Y-W%W', transaction_date)",
        "DATE_FORMAT(transaction_date, '%Y')": "strftime('%Y', transaction_date)",
        "DATE(created_at)=CURDATE()": "date(created_at)=date('now')",
    }
    for old, new in replacements.items():
        sql = sql.replace(old, new)
    return sql


def query(sql, params=(), one=False, commit=False):
    conn, driver = db()
    cur = conn.cursor(dictionary=True) if driver == "mysql" else conn.cursor()
    cur.execute(normalize_sql(sql, driver), params)
    if commit:
        conn.commit()
        data = cur.lastrowid
    else:
        rows = cur.fetchone() if one else cur.fetchall()
        if driver == "sqlite":
            if one:
                data = dict(rows) if rows else None
            else:
                data = [dict(row) for row in rows]
        else:
            data = rows
    cur.close()
    conn.close()
    return data


def create_token(user):
    payload = {
        "sub": str(user["id"]),
        "email": user["email"],
        "exp": datetime.utcnow() + timedelta(hours=8),
    }
    return jwt.encode(payload, JWT_SECRET, algorithm="HS256")


def current_user():
    token = session.get("jwt")
    if not token:
        return None
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=["HS256"])
    except jwt.PyJWTError:
        session.clear()
        return None
    return query("SELECT id, name, email, profile_photo, created_at FROM users WHERE id=%s", (payload["sub"],), one=True)


def login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not current_user():
            flash("Please login to continue.", "warning")
            return redirect(url_for("login"))
        return fn(*args, **kwargs)

    return wrapper


def csrf_token():
    token = session.get("csrf_token")
    if not token:
        token = os.urandom(24).hex()
        session["csrf_token"] = token
    return token


def google_login_enabled():
    return bool(os.getenv("GOOGLE_CLIENT_ID") and os.getenv("GOOGLE_CLIENT_SECRET"))


@app.context_processor
def inject_globals():
    return {
        "user": current_user(),
        "csrf_token": csrf_token(),
        "categories": CATEGORIES,
        "db_driver": ACTIVE_DB_DRIVER,
        "google_login_enabled": google_login_enabled(),
    }


def validate_csrf():
    form_token = request.form.get("csrf_token") or request.headers.get("X-CSRF-Token")
    if not form_token or form_token != session.get("csrf_token"):
        flash("Security token expired. Please try again.", "danger")
        return False
    return True


def parse_transaction_date(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if value is None:
        return None
    value = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(value, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def import_transactions_from_excel(file_storage, user_id):
    file_storage.stream.seek(0)
    workbook = load_workbook(file_storage.stream, data_only=True)
    sheet = workbook.active
    headers = [str(cell.value or "").strip().lower() for cell in sheet[1]]
    missing = [column for column in EXCEL_COLUMNS[:-1] if column not in headers]
    if missing:
        return 0, [f"Missing required columns: {', '.join(missing)}"]

    positions = {name: headers.index(name) for name in headers if name}
    imported = 0
    errors = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        title = str(row[positions["title"]] or "").strip()
        amount_raw = row[positions["amount"]]
        category = str(row[positions["category"]] or "").strip().title()
        tx_type = str(row[positions["type"]] or "").strip().title()
        tx_date = parse_transaction_date(row[positions["transaction_date"]])
        notes = str(row[positions.get("notes", -1)] or "").strip() if "notes" in positions else ""

        try:
            amount = float(amount_raw)
        except (TypeError, ValueError):
            amount = 0

        if not title or amount <= 0 or category not in CATEGORIES or tx_type not in ("Income", "Expense") or not tx_date:
            errors.append(f"Row {row_number}: check title, amount, category, type, and date.")
            continue

        query(
            """INSERT INTO transactions (user_id, title, amount, category, type, notes, transaction_date)
               VALUES (%s,%s,%s,%s,%s,%s,%s)""",
            (user_id, title, amount, category, tx_type, notes, tx_date),
            commit=True,
        )
        imported += 1
    return imported, errors


def month_window():
    today = date.today()
    return today.replace(day=1), today.year, today.month


def dashboard_data(user_id):
    start, year, month = month_window()
    today = date.today()
    income = query(
        "SELECT COALESCE(SUM(amount),0) total FROM transactions WHERE user_id=%s AND type='Income'",
        (user_id,),
        one=True,
    )["total"]
    expenses = query(
        "SELECT COALESCE(SUM(amount),0) total FROM transactions WHERE user_id=%s AND type='Expense'",
        (user_id,),
        one=True,
    )["total"]
    budget = query(
        "SELECT * FROM budgets WHERE user_id=%s AND month=%s AND year=%s",
        (user_id, month, year),
        one=True,
    )
    monthly_spend = query(
        """SELECT COALESCE(SUM(amount),0) total FROM transactions
           WHERE user_id=%s AND type='Expense' AND transaction_date >= %s""",
        (user_id, start),
        one=True,
    )["total"]
    category_rows = query(
        """SELECT category, COALESCE(SUM(amount),0) total FROM transactions
           WHERE user_id=%s AND type='Expense' GROUP BY category ORDER BY total DESC""",
        (user_id,),
    )
    monthly_rows = query(
        """SELECT DATE_FORMAT(transaction_date, '%Y-%m') label, COALESCE(SUM(amount),0) total
           FROM transactions WHERE user_id=%s AND type='Expense'
           GROUP BY label ORDER BY label LIMIT 12""",
        (user_id,),
    )
    recent = query(
        """SELECT id, title, amount, category, type, transaction_date
           FROM transactions WHERE user_id=%s ORDER BY transaction_date DESC, id DESC LIMIT 8""",
        (user_id,),
    )
    alert = make_budget_alert(user_id, budget, monthly_spend)
    days_elapsed = max(today.day, 1)
    days_in_month = calendar.monthrange(year, month)[1]
    predicted_spend = (float(monthly_spend or 0) / days_elapsed) * days_in_month
    top_category = category_rows[0]["category"] if category_rows else "Other"
    savings_target = max(float(income or 0) * 0.2, 0)
    return {
        "income": float(income or 0),
        "expenses": float(expenses or 0),
        "balance": float((income or 0) - (expenses or 0)),
        "budget": float(budget["monthly_budget"]) if budget else 0,
        "monthly_spend": float(monthly_spend or 0),
        "remaining_budget": float((budget["monthly_budget"] if budget else 0) - (monthly_spend or 0)),
        "category_rows": category_rows,
        "monthly_rows": monthly_rows,
        "recent": recent,
        "alert": alert,
        "insights": {
            "predicted_spend": round(predicted_spend, 2),
            "recommendation": f"Review {top_category} spending first; it is currently your largest expense category.",
            "savings": f"Try reserving ${savings_target:.2f} this month before discretionary purchases.",
        },
    }


def make_budget_alert(user_id, budget, spending):
    if not budget or not budget["monthly_budget"]:
        return None
    limit = float(budget["monthly_budget"])
    used = (float(spending or 0) / limit) * 100
    if used > 100:
        msg, alert_type = "You have exceeded your monthly budget.", "critical"
    elif used >= 100:
        msg, alert_type = "Budget Limit Reached!", "danger"
    elif used >= 80:
        msg, alert_type = "Warning! You have used 80% of your budget.", "warning"
    else:
        return None
    existing = query(
        "SELECT id FROM alerts WHERE user_id=%s AND message=%s AND DATE(created_at)=CURDATE()",
        (user_id, msg),
        one=True,
    )
    if not existing:
        query("INSERT INTO alerts (user_id, message, alert_type) VALUES (%s,%s,%s)", (user_id, msg, alert_type), commit=True)
    return {"message": msg, "type": alert_type, "used": round(used, 1)}


@app.route("/")
def landing():
    return render_template("landing.html")


@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        if not validate_csrf():
            return redirect(url_for("register"))
        name = request.form.get("name", "").strip()
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        confirm = request.form.get("confirm_password", "")
        if not name or not email or len(password) < 8 or password != confirm:
            flash("Use valid details and an 8+ character matching password.", "danger")
            return redirect(url_for("register"))
        try:
            user_id = query(
                "INSERT INTO users (name, email, password) VALUES (%s,%s,%s)",
                (name, email, generate_password_hash(password)),
                commit=True,
            )
        except (mysql.connector.IntegrityError, sqlite3.IntegrityError):
            flash("That email is already registered.", "danger")
            return redirect(url_for("register"))
        user = {"id": user_id, "email": email}
        session["jwt"] = create_token(user)
        flash("Welcome! Your finance workspace is ready.", "success")
        return redirect(url_for("dashboard"))
    return render_template("register.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        if not validate_csrf():
            return redirect(url_for("login"))
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        user = query("SELECT * FROM users WHERE email=%s", (email,), one=True)
        if not user or not check_password_hash(user["password"], password):
            flash("Invalid email or password.", "danger")
            return redirect(url_for("login"))
        session["jwt"] = create_token(user)
        flash("Logged in successfully.", "success")
        return redirect(url_for("dashboard"))
    return render_template("login.html")


@app.route("/login/google")
def google_login():
    if not google_login_enabled():
        flash("Google login is not configured yet. Add GOOGLE_CLIENT_ID and GOOGLE_CLIENT_SECRET to .env.", "warning")
        return redirect(url_for("login"))
    redirect_uri = url_for("google_callback", _external=True)
    return google.authorize_redirect(redirect_uri)


@app.route("/auth/google/callback")
def google_callback():
    if not google_login_enabled():
        flash("Google login is not configured yet.", "warning")
        return redirect(url_for("login"))
    try:
        token = google.authorize_access_token()
        profile = token.get("userinfo") or google.parse_id_token(token)
    except Exception:
        flash("Google login failed. Please try again.", "danger")
        return redirect(url_for("login"))

    email = (profile.get("email") or "").strip().lower()
    name = profile.get("name") or email.split("@")[0]
    if not email or not profile.get("email_verified", True):
        flash("Google account email could not be verified.", "danger")
        return redirect(url_for("login"))

    user = query("SELECT * FROM users WHERE email=%s", (email,), one=True)
    if not user:
        random_password = generate_password_hash(os.urandom(32).hex())
        user_id = query(
            "INSERT INTO users (name, email, password) VALUES (%s,%s,%s)",
            (name, email, random_password),
            commit=True,
        )
        user = {"id": user_id, "email": email}
    session["jwt"] = create_token(user)
    flash("Logged in with Google.", "success")
    return redirect(url_for("dashboard"))


@app.route("/logout")
def logout():
    session.clear()
    flash("You have been logged out.", "info")
    return redirect(url_for("landing"))


@app.route("/dashboard")
@login_required
def dashboard():
    data = dashboard_data(current_user()["id"])
    alerts = query("SELECT * FROM alerts WHERE user_id=%s ORDER BY created_at DESC LIMIT 6", (current_user()["id"],))
    return render_template("dashboard.html", data=data, alerts=alerts)


@app.route("/transactions", methods=["GET", "POST"])
@login_required
def transactions():
    user_id = current_user()["id"]
    if request.method == "POST":
        if not validate_csrf():
            return redirect(url_for("transactions"))
        title = request.form.get("title", "").strip()
        amount = request.form.get("amount", "0")
        category = request.form.get("category")
        tx_type = request.form.get("type")
        tx_date = request.form.get("transaction_date") or date.today().isoformat()
        notes = request.form.get("notes", "").strip()
        if not title or category not in CATEGORIES or tx_type not in ("Income", "Expense"):
            flash("Please complete all required transaction fields.", "danger")
            return redirect(url_for("transactions"))
        query(
            """INSERT INTO transactions (user_id, title, amount, category, type, notes, transaction_date)
               VALUES (%s,%s,%s,%s,%s,%s,%s)""",
            (user_id, title, amount, category, tx_type, notes, tx_date),
            commit=True,
        )
        dashboard_data(user_id)
        flash("Transaction saved.", "success")
        return redirect(url_for("dashboard"))
    search = request.args.get("search", "").strip()
    category = request.args.get("category", "")
    tx_type = request.args.get("type", "")
    sql = "SELECT * FROM transactions WHERE user_id=%s"
    params = [user_id]
    if search:
        sql += " AND (title LIKE %s OR notes LIKE %s)"
        params += [f"%{search}%", f"%{search}%"]
    if category in CATEGORIES:
        sql += " AND category=%s"
        params.append(category)
    if tx_type in ("Income", "Expense"):
        sql += " AND type=%s"
        params.append(tx_type)
    rows = query(sql + " ORDER BY transaction_date DESC, id DESC", tuple(params))
    return render_template("transactions.html", transactions=rows)


@app.route("/budget", methods=["GET", "POST"])
@login_required
def budget():
    user_id = current_user()["id"]
    start, year, month = month_window()
    if request.method == "POST":
        if not validate_csrf():
            return redirect(url_for("budget"))
        action = request.form.get("action")
        if action == "delete":
            query("DELETE FROM budgets WHERE user_id=%s AND month=%s AND year=%s", (user_id, month, year), commit=True)
            flash("Budget deleted.", "info")
        else:
            monthly_budget = request.form.get("monthly_budget", "0")
            existing = query("SELECT id FROM budgets WHERE user_id=%s AND month=%s AND year=%s", (user_id, month, year), one=True)
            if existing:
                query("UPDATE budgets SET monthly_budget=%s WHERE id=%s", (monthly_budget, existing["id"]), commit=True)
                flash("Budget updated.", "success")
            else:
                query(
                    "INSERT INTO budgets (user_id, monthly_budget, month, year) VALUES (%s,%s,%s,%s)",
                    (user_id, monthly_budget, month, year),
                    commit=True,
                )
                flash("Budget created.", "success")
        return redirect(url_for("budget"))
    data = dashboard_data(user_id)
    used_pct = round((data["monthly_spend"] / data["budget"]) * 100, 1) if data["budget"] else 0
    return render_template("budget.html", data=data, used_pct=used_pct)


@app.route("/reports")
@login_required
def reports():
    period = request.args.get("period", "monthly")
    return render_template("reports.html", period=period)


@app.route("/transactions/import", methods=["POST"])
@login_required
def import_transactions():
    if not validate_csrf():
        return redirect(url_for("reports"))
    upload = request.files.get("transaction_file")
    if not upload or not upload.filename:
        flash("Please choose an Excel file to upload.", "warning")
        return redirect(url_for("reports"))
    if not upload.filename.lower().endswith(".xlsx"):
        flash("Upload an .xlsx Excel file using the sample format.", "danger")
        return redirect(url_for("reports"))

    imported, errors = import_transactions_from_excel(upload, current_user()["id"])
    if imported:
        dashboard_data(current_user()["id"])
        flash(f"Imported {imported} transactions from Excel.", "success")
    if errors:
        preview = " ".join(errors[:3])
        extra = f" {len(errors) - 3} more rows skipped." if len(errors) > 3 else ""
        flash(preview + extra, "warning")
    if not imported and not errors:
        flash("No transaction rows were found in the Excel file.", "warning")
    return redirect(url_for("reports", period="monthly"))


@app.route("/sample/transactions.xlsx")
@login_required
def sample_transactions_excel():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Transactions"
    sheet.append(EXCEL_COLUMNS)
    sheet.append(["Groceries", 340.25, "Food", "Expense", date.today().replace(day=3).isoformat(), "Weekly grocery run"])
    sheet.append(["Monthly Salary", 4200, "Salary", "Income", date.today().replace(day=1).isoformat(), "Primary job"])
    sheet.append(["Metro Card", 65, "Transportation", "Expense", date.today().replace(day=5).isoformat(), "Commute pass"])
    for column in ("A", "B", "C", "D", "E", "F"):
        sheet.column_dimensions[column].width = 22
    out = io.BytesIO()
    workbook.save(out)
    out.seek(0)
    return Response(
        out.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=transaction-upload-sample.xlsx"},
    )


@app.route("/profile", methods=["GET", "POST"])
@login_required
def profile():
    user = current_user()
    if request.method == "POST":
        if not validate_csrf():
            return redirect(url_for("profile"))
        name = request.form.get("name", "").strip()
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        photo = request.files.get("profile_photo")
        filename = user.get("profile_photo")
        if photo and photo.filename:
            filename = secure_filename(photo.filename)
            photo.save(os.path.join(app.config["UPLOAD_FOLDER"], filename))
        if password:
            query(
                "UPDATE users SET name=%s, email=%s, password=%s, profile_photo=%s WHERE id=%s",
                (name, email, generate_password_hash(password), filename, user["id"]),
                commit=True,
            )
        else:
            query("UPDATE users SET name=%s, email=%s, profile_photo=%s WHERE id=%s", (name, email, filename, user["id"]), commit=True)
        flash("Profile updated.", "success")
        return redirect(url_for("profile"))
    return render_template("profile.html")


@app.route("/api/dashboard")
@login_required
def api_dashboard():
    data = dashboard_data(current_user()["id"])
    return jsonify(
        {
            "categoryLabels": [r["category"] for r in data["category_rows"]],
            "categoryValues": [float(r["total"]) for r in data["category_rows"]],
            "monthLabels": [r["label"] for r in data["monthly_rows"]],
            "monthValues": [float(r["total"]) for r in data["monthly_rows"]],
        }
    )


@app.route("/api/reports")
@login_required
def api_reports():
    user_id = current_user()["id"]
    period = request.args.get("period", "monthly")
    group_map = {
        "daily": "%Y-%m-%d",
        "weekly": "%x-W%v",
        "monthly": "%Y-%m",
        "yearly": "%Y",
    }
    fmt = group_map.get(period, "%Y-%m")
    trends = query(
        f"""SELECT DATE_FORMAT(transaction_date, '{fmt}') label, type, COALESCE(SUM(amount),0) total
            FROM transactions WHERE user_id=%s GROUP BY label, type ORDER BY label""",
        (user_id,),
    )
    categories = query(
        """SELECT category, COALESCE(SUM(amount),0) total FROM transactions
           WHERE user_id=%s AND type='Expense' GROUP BY category ORDER BY total DESC""",
        (user_id,),
    )
    return jsonify({"trends": trends, "categories": categories})


@app.route("/export/<kind>")
@login_required
def export_report(kind):
    user_id = current_user()["id"]
    rows = query(
        "SELECT transaction_date, title, category, type, amount FROM transactions WHERE user_id=%s ORDER BY transaction_date DESC",
        (user_id,),
    )
    if kind == "excel":
        out = io.StringIO()
        writer = csv.writer(out)
        writer.writerow(["Date", "Title", "Category", "Type", "Amount"])
        for row in rows:
            writer.writerow([row["transaction_date"], row["title"], row["category"], row["type"], row["amount"]])
        return Response(out.getvalue(), mimetype="text/csv", headers={"Content-Disposition": "attachment; filename=finance-report.csv"})
    text = "Personal Finance Report\n\n" + "\n".join(
        f"{r['transaction_date']} | {r['title']} | {r['category']} | {r['type']} | {r['amount']}" for r in rows
    )
    return Response(minimal_pdf(text), mimetype="application/pdf", headers={"Content-Disposition": "attachment; filename=finance-report.pdf"})


def minimal_pdf(text):
    safe = text.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")
    lines = safe.splitlines()[:42]
    stream = "BT /F1 12 Tf 50 780 Td " + " T* ".join(f"({line})" for line in lines) + " ET"
    objects = [
        "<< /Type /Catalog /Pages 2 0 R >>",
        "<< /Type /Pages /Kids [3 0 R] /Count 1 >>",
        "<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >>",
        "<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
        f"<< /Length {len(stream)} >>\nstream\n{stream}\nendstream",
    ]
    pdf = "%PDF-1.4\n"
    offsets = []
    for i, obj in enumerate(objects, 1):
        offsets.append(len(pdf))
        pdf += f"{i} 0 obj\n{obj}\nendobj\n"
    xref = len(pdf)
    pdf += "xref\n0 6\n0000000000 65535 f \n" + "".join(f"{o:010d} 00000 n \n" for o in offsets)
    pdf += f"trailer << /Root 1 0 R /Size 6 >>\nstartxref\n{xref}\n%%EOF"
    return pdf.encode("latin-1", "ignore")


if __name__ == "__main__":
    app.run(debug=True)
